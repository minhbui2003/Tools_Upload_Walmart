import os
import glob
from openpyxl import load_workbook

def replace_wung_in_excel(folder_path):
    files = glob.glob(os.path.join(folder_path, "**", "*.xlsx"), recursive=True)
    results = []
    
    for f in files:
        if "~$" in f:
            continue
        try:
            wb = load_workbook(f)
            modified = False
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "WUNG" in cell.value:
                            old_val = cell.value
                            new_val = old_val.replace("WUNG", "UNG")
                            cell.value = new_val
                            modified = True
                            results.append(f"File: {os.path.basename(f)} | Sheet: {sheet_name} | Cell: {cell.coordinate} | Changed: {old_val} -> {new_val}")
            
            if modified:
                wb.save(f)
        except Exception as e:
            pass
            
    return results

mapping_dir = r"d:\POD\Tools\POD_Marketplace\Upload-Toni\Source\Tools_Marketplace\mapping"
update_dir = r"d:\POD\Tools\POD_Marketplace\Upload-Toni\Source\Tools_Marketplace\update_File"
templates_dir = r"d:\POD\Tools\POD_Marketplace\Upload-Toni\Source\Tools_Marketplace\templates"

all_results = []
all_results.extend(replace_wung_in_excel(mapping_dir))
all_results.extend(replace_wung_in_excel(update_dir))
all_results.extend(replace_wung_in_excel(templates_dir))

if not all_results:
    print("No 'WUNG' found.")
else:
    for r in all_results:
        print(r)
