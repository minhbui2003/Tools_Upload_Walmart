"""
Walmart XLSX mapper: reads mapping config, merges getlinks + update data, writes final XLSX.
"""
import os
import csv
import shutil
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Border, Side

from core.config import SUPPORTED_SOURCE_TYPES
from core.utils import clean_text, is_enabled, normalize_columns
from core.sku import (
    get_sku_by_color,
    build_full_skuwa_from_row,
    get_base_prefix_for_row,
)
from core.xml_fixer import fix_walmart_xml


def normalize_date_output(value, mode):
    text = clean_text(value)

    if text == "":
        return ""

    if isinstance(value, datetime):
        if mode == "datetime":
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        if mode == "datetime":
            return datetime(value.year, value.month, value.day).strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")

    try:
        parsed = pd.to_datetime(text, errors="raise", dayfirst=False)

        if mode == "datetime":
            return parsed.strftime("%Y-%m-%d %H:%M:%S")

        return parsed.strftime("%Y-%m-%d")

    except Exception:
        return text


def format_value_for_seller_column(value, seller_excel_col, seller_col_name, expected_type=None):
    col = clean_text(seller_excel_col).upper()
    seller_name = clean_text(seller_col_name).lower()

    if col == "DR" or seller_name == "site start date":
        if seller_name == "site start date":
            return datetime.now().strftime("%Y-%m-%d 00:00:00")
        return normalize_date_output(value, mode="datetime")

    if col == "DS" or seller_name == "site end date":
        return normalize_date_output(value, mode="date")

    if pd.isna(value) or value is None:
        return ""
        
    cleaned = clean_text(value)
    
    # Bắt buộc ép kiểu chuỗi (String) cho cột Fulfillment Center ID
    if "fulfillment center id" in seller_name:
        return cleaned

    if expected_type == "string":
        return cleaned

    if expected_type == "number":
        try:
            if '.' in cleaned:
                return float(cleaned)
            else:
                return int(cleaned)
        except ValueError:
            return cleaned

    if isinstance(value, (int, float)):
        return value
        
    if len(cleaned) > 1 and cleaned.startswith("0") and not cleaned.startswith("0."):
        return cleaned
        
    try:
        if '.' in cleaned:
            return float(cleaned)
        else:
            return int(cleaned)
    except ValueError:
        return cleaned


def read_table_file(path):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(
            path,
            dtype=str,
            encoding="utf-8-sig",
            keep_default_na=False,
        ).fillna("")

    if ext in [".xlsx", ".xlsm", ".xls"]:
        return pd.read_excel(
            path,
            dtype=object,
            keep_default_na=False,
        ).fillna("")

    raise ValueError(f"Unsupported file type: {ext}")


def read_mapping_file(mapping_path):
    df = pd.read_excel(
        mapping_path,
        sheet_name="Mapping",
        dtype=object,
        keep_default_na=False,
    ).fillna("")

    required_cols = [
        "Enabled",
        "Seller Sheet",
        "Seller Header Row",
        "Seller Excel Col",
        "Seller Column",
        "Source Type",
        "Source Value",
        "Fixed Value",
    ]

    missing = [col for col in required_cols if col not in df.columns]

    if missing:
        raise ValueError(f"Mapping file missing columns: {missing}")

    return df


def get_seller_template_info(mapping_df):
    enabled_df = mapping_df[mapping_df["Enabled"].apply(is_enabled)]

    if enabled_df.empty:
        raise ValueError("No enabled rows found in Mapping sheet.")

    seller_sheet = clean_text(enabled_df.iloc[0].get("Seller Sheet"))
    header_row_raw = clean_text(enabled_df.iloc[0].get("Seller Header Row"))

    if not seller_sheet:
        seller_sheet = "Product Content And Site Exp"

    if not header_row_raw:
        header_row = 4
    else:
        header_row = int(float(header_row_raw))

    return seller_sheet, header_row


def prepare_update_df(update_df, sku_prefix, product_type, profile_manager, log_callback=None):
    """
    Tự tạo full SKUWA cho update_file nếu:
    - SKUWA đang là suffix như B00S / N00S / H00S / F00S
    - hoặc SKUWA đang trống nhưng có COLOR + SIZE
    - hoặc có SKU/Prefix riêng trong update file

    Ưu tiên:
    1. SKUWA full có UGNT/UNGT -> giữ nguyên
    2. SKUWA suffix -> ghép với SKU/Prefix/app prefix theo màu
    3. SKUWA trống -> tạo suffix từ COLOR + SIZE rồi ghép
    """
    if update_df is None or update_df.empty:
        return update_df

    update_df = normalize_columns(update_df.copy())

    if "SKUWA" not in update_df.columns:
        update_df["SKUWA"] = ""

    if "SKU" not in update_df.columns:
        update_df["SKU"] = ""

    update_df["SKUWA"] = update_df["SKUWA"].astype("object")
    update_df["SKU"] = update_df["SKU"].astype("object")

    for idx, row in update_df.iterrows():
        color = row.get("COLOR") if "COLOR" in update_df.columns else ""
        base_prefix = get_base_prefix_for_row(row, sku_prefix)

        adjusted_sku = get_sku_by_color(base_prefix, color, product_type, profile_manager)
        full_skuwa = build_full_skuwa_from_row(row, sku_prefix, product_type, profile_manager)

        update_df.at[idx, "SKU"] = adjusted_sku
        update_df.at[idx, "SKUWA"] = full_skuwa

    if log_callback:
        sample = update_df["SKUWA"].head(5).tolist()
        log_callback(f"Prepared update SKUWA sample: {sample}")

    if "COLOR" in update_df.columns and "SIZE" in update_df.columns:
        size_map = {
            "S": 1,
            "M": 2,
            "L": 3,
            "XL": 4,
            "2XL": 5,
            "XXL": 5,
            "3XL": 6,
            "XXXL": 6,
            "4XL": 7,
            "XXXXL": 7,
            "5XL": 8,
            "XXXXXL": 8,
        }
        update_df["_size_order"] = update_df["SIZE"].astype(str).str.strip().str.upper().map(lambda x: size_map.get(x, 99))
        update_df = update_df.sort_values(by=["_size_order", "COLOR"]).drop(columns=["_size_order"]).reset_index(drop=True)

    return update_df


def get_update_row_for_getlinks_row(getlinks_row, update_df):
    if update_df is None or update_df.empty:
        return None

    if "SKUWA" not in update_df.columns:
        if len(update_df) == 1:
            return update_df.iloc[0]
        return None

    skuwa = clean_text(getlinks_row.get("SKUWA"))

    if not skuwa:
        if len(update_df) == 1:
            return update_df.iloc[0]
        return None

    matched = update_df[update_df["SKUWA"].astype(str).str.strip() == skuwa]

    if not matched.empty:
        return matched.iloc[0]

    if len(update_df) == 1:
        return update_df.iloc[0]

    return None


def get_getlinks_row_for_update_row(update_row, getlinks_df):
    if getlinks_df is None or getlinks_df.empty:
        return None

    if "SKUWA" not in getlinks_df.columns:
        if len(getlinks_df) == 1:
            return getlinks_df.iloc[0]
        return None

    skuwa = clean_text(update_row.get("SKUWA"))

    if not skuwa:
        if len(getlinks_df) == 1:
            return getlinks_df.iloc[0]
        return None

    matched = getlinks_df[getlinks_df["SKUWA"].astype(str).str.strip() == skuwa]

    if not matched.empty:
        return matched.iloc[0]

    if len(getlinks_df) == 1:
        return getlinks_df.iloc[0]

    return None


def resolve_value(mapping_row, getlinks_row, update_row):
    source_type = clean_text(mapping_row.get("Source Type")).lower()
    source_value = clean_text(mapping_row.get("Source Value"))
    fixed_value = clean_text(mapping_row.get("Fixed Value"))

    if source_type == "getlinks":
        if not source_value:
            return "", "Missing Source Value for getlinks"

        if getlinks_row is None:
            return "", "No matching getlinks row"

        if source_value not in getlinks_row.index:
            return "", f"Getlinks column not found: {source_value}"

        return getlinks_row.get(source_value), ""

    if source_type == "update_file":
        if not source_value:
            return "", "Missing Source Value for update_file"

        if update_row is None:
            return "", "No matching update row"

        if source_value not in update_row.index:
            return "", f"Update file column not found: {source_value}"

        return update_row.get(source_value), ""

    if source_type == "fixed":
        return fixed_value, ""

    if source_type == "blank":
        return "", ""

    if source_type == "skip":
        return None, "Skipped"

    return "", f"Unsupported Source Type: {source_type}"


def clear_existing_data(ws, start_row):
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < start_row:
        return

    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def generate_walmart_xlsx_from_getlinks_df(
    seller_template_path,
    getlinks_df,
    mapping_path,
    update_file_path,
    output_folder,
    sku_prefix,
    product_type,
    profile_manager,
    log_callback=None,
):
    os.makedirs(output_folder, exist_ok=True)

    mapping_df = read_mapping_file(mapping_path)

    if update_file_path:
        update_df = read_table_file(update_file_path)
        update_df = prepare_update_df(update_df, sku_prefix, product_type, profile_manager, log_callback=log_callback)
    else:
        update_df = None

    seller_sheet, header_row = get_seller_template_info(mapping_df)
    data_start_row = header_row + 3

    enabled_mapping = mapping_df[mapping_df["Enabled"].apply(is_enabled)].copy()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = os.path.join(output_folder, f"final_walmart_upload_ready_{timestamp}.xlsx")
    log_csv = os.path.join(output_folder, f"mapper_log_{timestamp}.csv")
    
    import tempfile
    temp_dir = tempfile.mkdtemp()
    temp_xlsx = os.path.join(temp_dir, f"temp_{timestamp}.xlsx")
    temp_log_csv = os.path.join(temp_dir, f"temp_log_{timestamp}.csv")

    shutil.copy2(seller_template_path, temp_xlsx)

    wb = load_workbook(temp_xlsx)

    if seller_sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{seller_sheet}' not found in seller template. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[seller_sheet]

    clear_existing_data(ws, data_start_row)

    logs = []

    if log_callback:
        log_callback(f"Seller sheet: {seller_sheet}")
        log_callback(f"Seller header row: {header_row}")
        log_callback(f"Data start row: {data_start_row}")
        log_callback(f"Getlinks rows: {len(getlinks_df)}")
        if update_df is not None:
            log_callback(f"Update rows: {len(update_df)}")
        else:
            log_callback("No update file selected.")

    thin_border = Border(
        left=Side(style='thin', color='D4D4D4'),
        right=Side(style='thin', color='D4D4D4'),
        top=Side(style='thin', color='D4D4D4'),
        bottom=Side(style='thin', color='D4D4D4')
    )
    max_col = ws.max_column

    col_expected_types = {}
    type_row = header_row + 2
    if type_row <= ws.max_row:
        for c in range(1, max_col + 1):
            cell_val = ws.cell(row=type_row, column=c).value
            if cell_val and isinstance(cell_val, str):
                val_lower = cell_val.lower()
                if val_lower.startswith("alphanumeric") or val_lower.startswith("string") or val_lower.startswith("text") or val_lower.startswith("closed list"):
                    col_expected_types[c] = "string"
                elif val_lower.startswith("decimal") or val_lower.startswith("number") or val_lower.startswith("integer"):
                    col_expected_types[c] = "number"

    if update_df is not None and not update_df.empty:
        base_df = update_df
        iterate_by = "update"
    else:
        base_df = getlinks_df
        iterate_by = "getlinks"

    if log_callback:
        log_callback(f"Final row source: {iterate_by} ({len(base_df)} rows)")

        if iterate_by == "update" and "SKUWA" in update_df.columns and "SKUWA" in getlinks_df.columns:
            update_skuwas = set(update_df["SKUWA"].astype(str).str.strip())
            getlinks_skuwas = set(getlinks_df["SKUWA"].astype(str).str.strip())
            update_skuwas.discard("")
            getlinks_skuwas.discard("")

            unused_getlinks = getlinks_skuwas - update_skuwas
            unmatched_updates = update_skuwas - getlinks_skuwas

            if unused_getlinks:
                log_callback(f"Getlinks rows not in final because update file has no SKUWA match: {len(unused_getlinks)}")
            if unmatched_updates:
                log_callback(f"Update rows without getlinks match: {len(unmatched_updates)}")

    for idx in range(len(base_df)):
        row_data = base_df.iloc[idx]
        excel_row = data_start_row + idx

        if iterate_by == "update":
            update_row = row_data
            getlinks_row = get_getlinks_row_for_update_row(update_row, getlinks_df)
            skuwa_for_log = clean_text(update_row.get("SKUWA"))
        else:
            getlinks_row = row_data
            update_row = None
            skuwa_for_log = clean_text(getlinks_row.get("SKUWA"))

        # Apply borders to all columns in this row
        for c in range(1, max_col + 1):
            ws.cell(row=excel_row, column=c).border = thin_border

        for _, mapping_row in enabled_mapping.iterrows():
            seller_col_name = clean_text(mapping_row.get("Seller Column"))
            seller_excel_col = clean_text(mapping_row.get("Seller Excel Col"))
            source_type = clean_text(mapping_row.get("Source Type")).lower()
            source_value = clean_text(mapping_row.get("Source Value"))

            if not seller_excel_col:
                logs.append({
                    "Excel Row": excel_row,
                    "SKUWA": skuwa_for_log,
                    "Seller Column": seller_col_name,
                    "Source Type": source_type,
                    "Source Value": source_value,
                    "Status": "Error",
                    "Note": "Missing Seller Excel Col",
                })
                continue

            if source_type not in SUPPORTED_SOURCE_TYPES:
                logs.append({
                    "Excel Row": excel_row,
                    "SKUWA": skuwa_for_log,
                    "Seller Column": seller_col_name,
                    "Source Type": source_type,
                    "Source Value": source_value,
                    "Status": "Error",
                    "Note": f"Unsupported Source Type: {source_type}",
                })
                continue

            value, note = resolve_value(mapping_row, getlinks_row, update_row)

            if source_type == "skip":
                logs.append({
                    "Excel Row": excel_row,
                    "SKUWA": skuwa_for_log,
                    "Seller Column": seller_col_name,
                    "Source Type": source_type,
                    "Source Value": source_value,
                    "Status": "Skipped",
                    "Note": note,
                })
                continue

            try:
                col_index = column_index_from_string(seller_excel_col)
            except Exception:
                logs.append({
                    "Excel Row": excel_row,
                    "SKUWA": skuwa_for_log,
                    "Seller Column": seller_col_name,
                    "Source Type": source_type,
                    "Source Value": source_value,
                    "Status": "Error",
                    "Note": f"Invalid Seller Excel Col: {seller_excel_col}",
                })
                continue

            expected_type = col_expected_types.get(col_index)

            formatted_value = format_value_for_seller_column(
                value=value,
                seller_excel_col=seller_excel_col,
                seller_col_name=seller_col_name,
                expected_type=expected_type,
            )

            cell = ws.cell(row=excel_row, column=col_index)

            if expected_type == "string" or clean_text(seller_excel_col).upper() in ["DR", "DS"] or seller_col_name.lower() in ["site start date", "site end date", "fulfillment center id"]:
                cell.number_format = "@"

            if formatted_value == "":
                cell.value = None
            else:
                cell.value = formatted_value

            status = "OK"
            if note and note != "Skipped":
                status = "Warning"

            logs.append({
                "Excel Row": excel_row,
                "SKUWA": skuwa_for_log,
                "Seller Column": seller_col_name,
                "Source Type": source_type,
                "Source Value": source_value,
                "Status": status,
                "Note": note,
            })

    wb.save(temp_xlsx)

    with open(temp_log_csv, "w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "Excel Row",
            "SKUWA",
            "Seller Column",
            "Source Type",
            "Source Value",
            "Status",
            "Note",
        ]

        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for item in logs:
            writer.writerow(item)
            
    shutil.copy2(temp_xlsx, output_xlsx)
    shutil.copy2(temp_log_csv, log_csv)
    
    # Fix openpyxl internal XML quirks (inlineStr -> sharedStrings) for Walmart
    fix_success = fix_walmart_xml(output_xlsx, log_callback)
    
    if not fix_success:
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            wb_excel = excel.Workbooks.Open(os.path.abspath(output_xlsx))
            wb_excel.Save()
            wb_excel.Close(SaveChanges=True)
            excel.Quit()
            if log_callback:
                log_callback("File optimized successfully using Microsoft Excel.")
        except Exception as e:
            if log_callback:
                log_callback(f"Excel optimization skipped: {e}")
    
    try:
        shutil.rmtree(temp_dir)
    except Exception:
        pass

    return output_xlsx, log_csv
