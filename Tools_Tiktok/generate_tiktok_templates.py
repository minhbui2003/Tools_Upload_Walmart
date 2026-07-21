import os
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.tiktok_generator import (
    PRODUCT_PROFILES,
    TIKTOK_DEFAULTS,
    create_tiktok_image_mixing_template,
    create_tiktok_update_template,
)


ROOT = Path(__file__).resolve().parent
TEMPLATE_DIR = ROOT / "templates"
TEMPLATE_DIR.mkdir(exist_ok=True)

PROFILE_TEMPLATE_PATHS = {
    "mens_tshirt": ROOT / "Tiktoksellercenter_Menswear & Underwear_Template.xlsx",
    "womens_tshirt": ROOT / "Files" / "Tiktoksellercenter_160726HN6_20260714_Women's blank template.xlsx",
}

FALLBACK_TEMPLATE_COLUMNS = {
    "mens_tshirt": [
        ("category", "Category", "Mandatory"),
        ("brand", "Brand", "Optional"),
        ("product_name", "Product name", "Mandatory"),
        ("product_description", "Product description", "Mandatory"),
        ("main_image", "Main image", "Mandatory"),
        ("image_2", "Image 2", "Optional"),
        ("image_3", "Image 3", "Optional"),
        ("image_4", "Image 4", "Optional"),
        ("image_5", "Image 5", "Optional"),
        ("image_6", "Image 6", "Optional"),
        ("image_7", "Image 7", "Optional"),
        ("image_8", "Product Image 8", "Optional"),
        ("image_9", "Product Image 9", "Optional"),
        ("gtin_type", "Identifier Code Type", "Optional"),
        ("gtin_code", "Identifier Code", "Optional"),
        ("property_name_1", "Primary variation name (theme)", "Conditionally mandatory "),
        ("property_value_1", "Primary variation value (option)", "Conditionally mandatory "),
        ("property_1_image", "Primary variation image 1", "Optional"),
        ("property_1_image_2", "Primary variation image 2", "Optional"),
        ("property_1_image_3", "Primary variation image 3", "Optional"),
        ("property_1_image_4", "Primary variation image 4", "Optional"),
        ("property_1_image_5", "Primary variation image 5", "Optional"),
        ("property_1_image_6", "Primary variation image 6", "Optional"),
        ("property_1_image_7", "Primary variation image 7", "Optional"),
        ("property_1_image_8", "Primary variation image 8", "Optional"),
        ("property_1_image_9", "Primary variation image 9", "Optional"),
        ("property_name_2", "Secondary variation name (theme)", "Conditionally mandatory "),
        ("property_value_2", "Secondary variation value (option)", "Optional"),
        ("parcel_weight", "Package weight(lb)", "Mandatory"),
        ("parcel_length", "Package length(inch)", "Conditionally mandatory "),
        ("parcel_width", "Package width(inch)", "Conditionally mandatory "),
        ("parcel_height", "Package height(inch)", "Conditionally mandatory "),
        ("delivery", "Delivery options", "Optional"),
        ("price", "Retail Price (Local Currency)", "Mandatory"),
        ("list_price", "List price (Local currency)", "Optional"),
        ("quantity", "Quantity", "Mandatory"),
        ("seller_sku", "Seller SKU", "Optional"),
        ("size_chart", "Size Chart", "Conditionally mandatory "),
        ("special_product_listing_type", "Auction product", "Optional"),
        ("auction_starting_price", "Starting bid", "Optional"),
        ("product_property/100157", "Materials", "Conditionally mandatory "),
        ("product_property/100198", "Pattern", "Conditionally mandatory "),
        ("product_property/100393", "Neckline", "Conditionally mandatory "),
        ("product_property/100395", "Sleeve Length", "Conditionally mandatory "),
        ("product_property/100397", "Season", "Conditionally mandatory "),
        ("product_property/100398", "Style", "Conditionally mandatory "),
        ("product_property/100399", "Fit", "Conditionally mandatory "),
        ("product_property/100400", "Stretch", "Conditionally mandatory "),
        ("product_property/100401", "Washing Instructions", "Conditionally mandatory "),
        ("product_property/100403", "Waist Height", "Conditionally mandatory "),
        ("product_property/101619", "Dangerous Goods Or Hazardous Materials", "Conditionally mandatory "),
        ("product_property/102418", "Organic Textile", "Conditionally mandatory "),
        ("product_property/101395", "CA Prop 65: Repro. Chems", "Conditionally mandatory "),
        ("product_property/101398", "Reprotoxic Chemicals", "Conditionally mandatory "),
        ("product_property/101400", "CA Prop 65: Carcinogens", "Conditionally mandatory "),
        ("product_property/101397", "Carcinogen", "Conditionally mandatory "),
        (
            "qualification/8647636475739801353",
            "Safety Data Sheet (SDS) for other dangerous goods or hazardous materials",
            "Conditionally mandatory ",
        ),
        ("aimed_product_status", "Product Status", "Optional"),
    ],
    "womens_tshirt": [
        ("category", "Category", "Mandatory"),
        ("brand", "Brand", "Optional"),
        ("product_name", "Product name", "Mandatory"),
        ("product_description", "Product description", "Mandatory"),
        ("main_image", "Main image", "Mandatory"),
        ("image_2", "Image 2", "Optional"),
        ("image_3", "Image 3", "Optional"),
        ("image_4", "Image 4", "Optional"),
        ("image_5", "Image 5", "Optional"),
        ("image_6", "Image 6", "Optional"),
        ("image_7", "Image 7", "Optional"),
        ("image_8", "Product Image 8", "Optional"),
        ("image_9", "Product Image 9", "Optional"),
        ("gtin_type", "Identifier Code Type", "Optional"),
        ("gtin_code", "Identifier Code", "Optional"),
        ("property_name_1", "Primary variation name (theme)", "Conditionally mandatory "),
        ("property_value_1", "Primary variation value (option)", "Conditionally mandatory "),
        ("property_1_image", "Primary variation image 1", "Optional"),
        ("property_1_image_2", "Primary variation image 2", "Optional"),
        ("property_1_image_3", "Primary variation image 3", "Optional"),
        ("property_1_image_4", "Primary variation image 4", "Optional"),
        ("property_1_image_5", "Primary variation image 5", "Optional"),
        ("property_1_image_6", "Primary variation image 6", "Optional"),
        ("property_1_image_7", "Primary variation image 7", "Optional"),
        ("property_1_image_8", "Primary variation image 8", "Optional"),
        ("property_1_image_9", "Primary variation image 9", "Optional"),
        ("property_name_2", "Secondary variation name (theme)", "Conditionally mandatory "),
        ("property_value_2", "Secondary variation value (option)", "Optional"),
        ("parcel_weight", "Package weight(lb)", "Mandatory"),
        ("parcel_length", "Package length(inch)", "Conditionally mandatory "),
        ("parcel_width", "Package width(inch)", "Conditionally mandatory "),
        ("parcel_height", "Package height(inch)", "Conditionally mandatory "),
        ("delivery", "Delivery options", "Optional"),
        ("price", "Retail Price (Local Currency)", "Mandatory"),
        ("list_price", "List price (Local currency)", "Optional"),
        ("quantity", "Quantity", "Mandatory"),
        ("seller_sku", "Seller SKU", "Optional"),
        ("size_chart", "Size Chart", "Conditionally mandatory "),
        ("special_product_listing_type", "Auction product", "Optional"),
        ("auction_starting_price", "Starting bid", "Optional"),
        ("product_property/100157", "Materials", "Conditionally mandatory "),
        ("product_property/100198", "Pattern", "Conditionally mandatory "),
        ("product_property/100393", "Neckline", "Conditionally mandatory "),
        ("product_property/100395", "Sleeve Length", "Conditionally mandatory "),
        ("product_property/100396", "Sleeve Type", "Conditionally mandatory "),
        ("product_property/100397", "Season", "Conditionally mandatory "),
        ("product_property/100398", "Style", "Conditionally mandatory "),
        ("product_property/100399", "Fit", "Conditionally mandatory "),
        ("product_property/100400", "Stretch", "Conditionally mandatory "),
        ("product_property/100401", "Washing Instructions", "Conditionally mandatory "),
        ("product_property/100403", "Waist Height", "Conditionally mandatory "),
        ("product_property/101619", "Dangerous Goods Or Hazardous Materials", "Conditionally mandatory "),
        ("product_property/102418", "Organic Textile", "Conditionally mandatory "),
        ("product_property/101395", "CA Prop 65: Repro. Chems", "Conditionally mandatory "),
        ("product_property/101398", "Reprotoxic Chemicals", "Conditionally mandatory "),
        ("product_property/101400", "CA Prop 65: Carcinogens", "Conditionally mandatory "),
        ("product_property/101397", "Carcinogen", "Conditionally mandatory "),
        (
            "qualification/8647636475739801353",
            "Safety Data Sheet (SDS) for other dangerous goods or hazardous materials",
            "Conditionally mandatory ",
        ),
        ("aimed_product_status", "Product Status", "Optional"),
    ],
}

PROFILE_OUTPUTS = {
    "mens_tshirt": [
        TEMPLATE_DIR / "Tiktok_Mapping_Mens_Tshirt.xlsx",
        TEMPLATE_DIR / "Tiktok_Mapping_Mens_Tshirt_Default.xlsx",
        TEMPLATE_DIR / "Tiktok_Mapping_Mens_Tshirt_Full.xlsx",
    ],
    "womens_tshirt": [
        TEMPLATE_DIR / "Tiktok_Mapping_Womens_Tshirt.xlsx",
        TEMPLATE_DIR / "Tiktok_Mapping_Womens_Tshirt_Default.xlsx",
        TEMPLATE_DIR / "Tiktok_Mapping_Womens_Tshirt_Full.xlsx",
    ],
}

UPDATE_TEMPLATE_OUTPUTS = {
    "mens_tshirt": TEMPLATE_DIR / "Tiktok_Update_Mens_Tshirt.xlsx",
    "womens_tshirt": TEMPLATE_DIR / "Tiktok_Update_Womens_Tshirt.xlsx",
}

IMAGE_MIXING_OUTPUTS = {
    "mens_tshirt": TEMPLATE_DIR / "Tiktok_Image_Mixing_Mens.xlsx",
    "womens_tshirt": TEMPLATE_DIR / "Tiktok_Image_Mixing_Womens.xlsx",
}

NUMERIC_COLUMNS = {
    "parcel_weight",
    "parcel_length",
    "parcel_width",
    "parcel_height",
    "price",
    "list_price",
    "quantity",
    "auction_starting_price",
}

FIXED_BLANK_COLUMNS = {
    "gtin_code",
    "list_price",
    "special_product_listing_type",
    "auction_starting_price",
    "property_1_image_2",
    "property_1_image_3",
    "property_1_image_4",
    "property_1_image_5",
    "property_1_image_6",
    "property_1_image_7",
    "property_1_image_8",
    "property_1_image_9",
    "product_property/100403",
    "product_property/101398",
    "product_property/101397",
    "qualification/8647636475739801353",
}

FIXED_VALUES = {
    "gtin_type": "GTIN",
    "property_name_1": "Color",
    "property_name_2": "Size",
    "delivery": "",
    "aimed_product_status": "Active(1)",
}

UPDATE_SOURCE_VALUES = {
    "brand": "Brand",
    "product_name": "Product Name",
    "product_description": "Product Description",
    "price": "Price",
    "quantity": "Quantity",
    "size_chart": "Size Chart",
}

AUTO_SOURCE_VALUES = {
    "property_value_1": "property_value_1",
    "property_1_image": "property_1_image",
    "property_value_2": "property_value_2",
    "seller_sku": "seller_sku",
    "image_3": "image_3",
    "image_4": "image_4",
    "image_5": "image_5",
    "image_6": "image_6",
}

GETLINK_SOURCE_VALUES = {
    "main_image": "MK1",
    "image_2": "MK2",
}

STRICT_UPDATE_IMAGE_VALUES = {
    "image_7": "Image 8",
    "image_8": "Image 9",
}


def expected_type(column_key):
    return "number" if column_key in NUMERIC_COLUMNS else "string"


def rule_for_column(profile_key, column_key):
    profile = PRODUCT_PROFILES[profile_key]

    if column_key == "category":
        return "fixed", "", profile["category"], "TikTok category for this profile."

    if column_key in UPDATE_SOURCE_VALUES:
        source_value = UPDATE_SOURCE_VALUES[column_key]
        return "auto", source_value, "", "Uses update file first, then the tool default."

    if column_key in GETLINK_SOURCE_VALUES:
        return "getlinks", GETLINK_SOURCE_VALUES[column_key], "", "Resolved from the selected image folder."

    if column_key in AUTO_SOURCE_VALUES:
        return "auto", AUTO_SOURCE_VALUES[column_key], "", "Generated by the tool; image mixing can change image order."

    if column_key in STRICT_UPDATE_IMAGE_VALUES:
        return "update_file", STRICT_UPDATE_IMAGE_VALUES[column_key], "", "Strictly reads this update-file image column."

    if column_key == "image_9":
        return "blank", "", "", "Blank by default; change to update_file if needed."

    if column_key in FIXED_VALUES:
        return "fixed", "", FIXED_VALUES[column_key], "Fixed TikTok value."

    if column_key in TIKTOK_DEFAULTS:
        return "fixed", "", TIKTOK_DEFAULTS[column_key], "Fixed TikTok value."

    if column_key in FIXED_BLANK_COLUMNS:
        return "blank", "", "", "Optional blank column."

    return "skip", "", "", "Not managed by the tool."


def build_mapping_rows(profile_key, template_path):
    columns = []
    if Path(template_path).exists():
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported.*")
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
            wb = load_workbook(template_path, read_only=False, data_only=True)
        ws = wb["Template"]
        for col in range(1, ws.max_column + 1):
            column_key = ws.cell(row=1, column=col).value
            if column_key:
                columns.append(
                    (
                        str(column_key).strip(),
                        ws.cell(row=3, column=col).value or "",
                        ws.cell(row=4, column=col).value or "",
                    )
                )
    else:
        columns = FALLBACK_TEMPLATE_COLUMNS[profile_key]

    rows = []

    for col, (column_key, seller_label, requirement) in enumerate(columns, start=1):
        source_type, source_value, fixed_value, notes = rule_for_column(profile_key, column_key)
        rows.append(
            {
                "Enabled": "TRUE",
                "Seller Sheet": "Template",
                "Seller Header Row": 3,
                "Seller Excel Col": get_column_letter(col),
                "Seller Column": column_key,
                "Seller Label": seller_label,
                "Requirement": requirement,
                "Source Type": source_type,
                "Source Value": source_value,
                "Fixed Value": fixed_value,
                "Expected Type": expected_type(column_key),
                "Notes": notes,
            }
        )

    return rows


def write_mapping(profile_key):
    template_path = PROFILE_TEMPLATE_PATHS[profile_key]
    rows = build_mapping_rows(profile_key, template_path)
    df = pd.DataFrame(rows)

    for output_path in PROFILE_OUTPUTS[profile_key]:
        df.to_excel(output_path, sheet_name="Mapping", index=False)
        print(f"Generated {output_path.relative_to(ROOT)}")


def main():
    os.chdir(ROOT)

    for profile_key, output_path in UPDATE_TEMPLATE_OUTPUTS.items():
        create_tiktok_update_template(str(output_path), profile_key=profile_key)
        print(f"Generated {output_path.relative_to(ROOT)}")

    for profile_key, output_path in IMAGE_MIXING_OUTPUTS.items():
        create_tiktok_image_mixing_template(str(output_path), profile_key=profile_key)
        print(f"Generated {output_path.relative_to(ROOT)}")

    for profile_key in PROFILE_OUTPUTS:
        write_mapping(profile_key)

    print("Template generation finished.")


if __name__ == "__main__":
    main()
