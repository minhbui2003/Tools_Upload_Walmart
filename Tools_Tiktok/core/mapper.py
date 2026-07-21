import os
import csv
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Border, Side

from core.utils import clean_text, is_enabled, normalize_columns
from core.sku import get_sku_by_color, build_full_skuwa_from_row, get_base_prefix_for_row


def format_value_for_tiktok(value, expected_type="string"):
    if pd.isna(value) or value is None:
        return ""
    
    cleaned = clean_text(value)

    if expected_type == "string":
        return cleaned

    if expected_type == "number":
        try:
            if '.' in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return cleaned

    return cleaned


def read_table_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path, dtype=str, encoding="utf-8-sig", keep_default_na=False).fillna("")
    if ext in [".xlsx", ".xlsm", ".xls"]:
        return pd.read_excel(path, dtype=object, keep_default_na=False).fillna("")
    raise ValueError(f"Unsupported file type: {ext}")


def prepare_update_df(update_df, sku_prefix, product_type, profile_manager, log_callback=None):
    if update_df is None or update_df.empty:
        return update_df

    update_df = normalize_columns(update_df.copy())

    if "SKUWA" not in update_df.columns:
        update_df["SKUWA"] = ""

    if "SKU" not in update_df.columns:
        update_df["SKU"] = ""

    for idx, row in update_df.iterrows():
        color = row.get("COLOR", "")
        base_prefix = get_base_prefix_for_row(row, sku_prefix)
        adjusted_sku = get_sku_by_color(base_prefix, color, product_type, profile_manager)
        full_skuwa = build_full_skuwa_from_row(row, sku_prefix, product_type, profile_manager)

        update_df.at[idx, "SKU"] = adjusted_sku
        update_df.at[idx, "SKUWA"] = full_skuwa

    if "COLOR" in update_df.columns and "SIZE" in update_df.columns:
        size_map = {"S": 1, "M": 2, "L": 3, "XL": 4, "2XL": 5, "XXL": 5, "3XL": 6, "XXXL": 6, "4XL": 7, "XXXXL": 7, "5XL": 8, "XXXXXL": 8}
        update_df["_size_order"] = update_df["SIZE"].astype(str).str.strip().str.upper().map(lambda x: size_map.get(x, 99))
        update_df = update_df.sort_values(by=["_size_order", "COLOR"]).drop(columns=["_size_order"]).reset_index(drop=True)

    return update_df


def get_matching_row(source_row, target_df):
    if target_df is None or target_df.empty:
        return None

    if "SKUWA" not in target_df.columns:
        return target_df.iloc[0] if len(target_df) == 1 else None

    skuwa = clean_text(source_row.get("SKUWA"))
    if not skuwa:
        return target_df.iloc[0] if len(target_df) == 1 else None

    matched = target_df[target_df["SKUWA"].astype(str).str.strip() == skuwa]
    if not matched.empty:
        return matched.iloc[0]
    
    return target_df.iloc[0] if len(target_df) == 1 else None


def generate_tiktok_xlsx_from_getlinks_df(
    seller_template_path,
    getlinks_df,
    mapping_path,
    update_file_path,
    output_folder,
    sku_prefix,
    product_type,
    profile_manager,
    log_callback=None,
):
    os.makedirs(output_folder, exist_ok=True)

    mapping_df = pd.read_excel(mapping_path, sheet_name="Mapping", dtype=object, keep_default_na=False).fillna("")
    
    if update_file_path:
        update_df = read_table_file(update_file_path)
        update_df = prepare_update_df(update_df, sku_prefix, product_type, profile_manager, log_callback)
    else:
        update_df = None

    enabled_mapping = mapping_df[mapping_df["Enabled"].apply(is_enabled)].copy()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = os.path.join(output_folder, f"final_tiktok_upload_ready_{timestamp}.xlsx")
    log_csv = os.path.join(output_folder, f"mapper_log_{timestamp}.csv")

    shutil.copy2(seller_template_path, output_xlsx)
    wb = load_workbook(output_xlsx)
    
    seller_sheet = "Template"
    if seller_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{seller_sheet}' not found in TikTok template.")
    
    ws = wb[seller_sheet]
    data_start_row = 7  # TikTok template data starts at row 7
    
    # Xoá dữ liệu cũ nếu có từ dòng 7 trở đi
    if ws.max_row >= data_start_row:
        for r in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
            for cell in r:
                cell.value = None

    logs = []
    
    thin_border = Border(left=Side(style='thin', color='D4D4D4'), right=Side(style='thin', color='D4D4D4'),
                         top=Side(style='thin', color='D4D4D4'), bottom=Side(style='thin', color='D4D4D4'))
    
    max_col = ws.max_column
    
    base_df = update_df if (update_df is not None and not update_df.empty) else getlinks_df
    iterate_by = "update" if (update_df is not None and not update_df.empty) else "getlinks"

    for idx in range(len(base_df)):
        row_data = base_df.iloc[idx]
        excel_row = data_start_row + idx

        if iterate_by == "update":
            update_row = row_data
            getlinks_row = get_matching_row(update_row, getlinks_df)
            skuwa_for_log = clean_text(update_row.get("SKUWA"))
        else:
            getlinks_row = row_data
            update_row = None
            skuwa_for_log = clean_text(getlinks_row.get("SKUWA"))

        for c in range(1, max_col + 1):
            ws.cell(row=excel_row, column=c).border = thin_border

        for _, mapping_row in enabled_mapping.iterrows():
            seller_excel_col = clean_text(mapping_row.get("Seller Excel Col"))
            seller_col_name = clean_text(mapping_row.get("Seller Column"))
            source_type = clean_text(mapping_row.get("Source Type")).lower()
            source_value = clean_text(mapping_row.get("Source Value"))
            expected_type = clean_text(mapping_row.get("Expected Type", "string")).lower()

            if not seller_excel_col:
                continue

            try:
                col_index = column_index_from_string(seller_excel_col)
            except Exception:
                continue

            value, note = "", ""
            if source_type == "getlinks" and getlinks_row is not None and source_value in getlinks_row.index:
                value = getlinks_row.get(source_value)
            elif source_type == "update_file" and update_row is not None and source_value in update_row.index:
                value = update_row.get(source_value)
            elif source_type == "fixed":
                value = clean_text(mapping_row.get("Fixed Value"))
            elif source_type == "blank":
                value = ""

            formatted_value = format_value_for_tiktok(value, expected_type)
            cell = ws.cell(row=excel_row, column=col_index)
            
            if expected_type == "string":
                cell.number_format = "@"
                
            cell.value = formatted_value

            logs.append({"Excel Row": excel_row, "SKUWA": skuwa_for_log, "Column": seller_col_name, "Status": "OK", "Value": formatted_value})

    wb.save(output_xlsx)
    
    with open(log_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["Excel Row", "SKUWA", "Column", "Status", "Value"])
        writer.writeheader()
        writer.writerows(logs)

    return output_xlsx, log_csv
