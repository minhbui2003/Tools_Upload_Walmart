"""
Generate TikTok Seller Center upload workbooks from a blank template.
"""
import csv
import os
import random
import shutil
import warnings
from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import column_index_from_string

from core.image import ImageResolver, normalize_asset_key
from core.utils import clean_text, is_enabled


DATA_START_ROW = 7
TEMPLATE_SHEET = "Template"

SIZE_DEFINITIONS = [
    {"input": "S", "label": "Small", "sku_code": "00S"},
    {"input": "M", "label": "Medium", "sku_code": "00M"},
    {"input": "L", "label": "Large", "sku_code": "00L"},
    {"input": "XL", "label": "X Large", "sku_code": "0XL"},
    {"input": "2XL", "label": "XX Large", "sku_code": "2XL"},
    {"input": "3XL", "label": "3XL", "sku_code": "3XL"},
    {"input": "4XL", "label": "4XL", "sku_code": "4XL"},
    {"input": "5XL", "label": "5XL", "sku_code": "5XL"},
]

PRODUCT_PROFILES = {
    "mens_tshirt": {
        "label": "T-Shirt - Men",
        "category": "Men's Tops/T-shirts/Men's Short-sleeve T-shirts",
        "default_name": "Men's Short-sleeve Graphic T-Shirt",
        "colors": [
            {"name": "Black", "sku_root": "UNGT1", "sku_letter": "B", "image_suffix": "1", "aliases": ["Black"]},
            {"name": "Navy", "sku_root": "UNGT1", "sku_letter": "N", "image_suffix": "2", "aliases": ["Navy"]},
            {"name": "Sport Grey", "sku_root": "UNGT4", "sku_letter": "H", "image_suffix": "3", "aliases": ["Sport Grey", "Grey", "SportGray", "Gray"]},
            {"name": "Military Green", "sku_root": "UNGT3", "sku_letter": "F", "image_suffix": "4", "aliases": ["Military Green", "Green"]},
        ],
    },
    "womens_tshirt": {
        "label": "T-Shirt - Women",
        "category": "Women's Tops/Women's T-shirts",
        "default_name": "Women's Short-sleeve Graphic T-Shirt",
        "colors": [
            {"name": "Black", "sku_root": "UNGT1", "sku_letter": "B", "image_suffix": "1", "aliases": ["Black"]},
            {"name": "Sand", "sku_root": "UNGT4", "sku_letter": "Y", "image_suffix": "2", "aliases": ["Sand"]},
            {"name": "Sport Grey", "sku_root": "UNGT4", "sku_letter": "H", "image_suffix": "3", "aliases": ["Sport Grey", "Grey", "SportGray", "Gray"]},
            {"name": "White", "sku_root": "UNGT1", "sku_letter": "W", "image_suffix": "4", "aliases": ["White"]},
        ],
    },
}

DEFAULT_DESCRIPTION = (
    "Premium graphic T-shirt made for everyday comfort.\n\n"
    "Features\n"
    "Soft, breathable fabric with a classic fit.\n"
    "High-quality, fade-resistant print designed to stay vibrant.\n"
    "Great for casual wear, birthdays, holidays, events, and gifts.\n\n"
    "Care Instructions\n"
    "Machine wash cold, inside out.\n"
    "Tumble dry low or hang dry.\n"
    "Do not bleach.\n"
    "Do not iron directly on the printed design."
)

TIKTOK_DEFAULTS = {
    "brand": "No brand",
    "parcel_weight": 0.5,
    "parcel_length": 12.0,
    "parcel_width": 9.0,
    "parcel_height": 1.0,
    "price": 19.99,
    "quantity": 100,
    "size_chart": "7660455797738030862",
    "product_property/100157": "Cotton",
    "product_property/100198": "Graphic",
    "product_property/100393": "Crew Neck",
    "product_property/100395": "Short Sleeve",
    "product_property/100396": "Short",
    "product_property/100397": "All Seasons",
    "product_property/100398": "Casual",
    "product_property/100399": "Regular",
    "product_property/100400": "Slight Stretch",
    "product_property/100401": "Machine Washable",
    "product_property/101619": "No",
    "product_property/102418": "No",
    "product_property/101395": "No",
    "product_property/101400": "No",
    "aimed_product_status": "Active(1)",
}

UPDATE_ALIASES = {
    "SKUWA": ["SKUWA", "Seller SKU", "seller_sku"],
    "COLOR": ["COLOR", "Color", "primary_variation_value", "Primary variation value (option)"],
    "SIZE": ["SIZE", "Size", "secondary_variation_value", "Secondary variation value (option)"],
    "TIKTOK_SIZE": ["TikTok Size", "TikTok size", "TIKTOK SIZE", "Size Label"],
    "PRODUCT_NAME": ["Product Name", "product_name", "PRODUCT NAME", "Title", "Name"],
    "PRODUCT_DESCRIPTION": ["Product Description", "product_description", "PRODUCT DESCRIPTION", "Description"],
    "MAIN_IMAGE": ["Main Image", "Main Image URL", "main_image", "main_image_url"],
    "IMAGE_2": ["Image 2", "Image 2 URL", "Product Image 2", "image_2", "image_2_url"],
    "IMAGE_3": ["Image 3", "Image 3 URL", "Product Image 3", "image_3", "image_3_url"],
    "IMAGE_4": ["Image 4", "Image 4 URL", "Product Image 4", "image_4", "image_4_url"],
    "IMAGE_5": ["Image 5", "Image 5 URL", "Product Image 5", "image_5", "image_5_url"],
    "IMAGE_6": ["Image 6", "Image 6 URL", "Product Image 6", "image_6", "image_6_url"],
    "IMAGE_7": ["Image 7", "Image 7 URL", "Product Image 7", "image_7", "image_7_url"],
    "IMAGE_8": ["Image 8", "Image 8 URL", "Product Image 8", "image_8", "image_8_url"],
    "IMAGE_9": ["Image 9", "Product Image 9", "Image 9 URL", "image_9", "image_9_url"],
    "PRICE": ["Price", "Retail Price", "Retail Price (Local Currency)", "price"],
    "QUANTITY": ["Quantity", "Qty", "quantity", "Inventory"],
    "SIZE_CHART": ["Size Chart", "size_chart"],
    "BRAND": ["Brand", "brand"],
}

IMAGE_MIX_OUTPUTS = [
    ("main_image", "Main Image Source", "MK1"),
    ("image_2", "Image 2 Source", "MK2"),
    ("image_3", "Image 3 Source", "MK3"),
    ("image_4", "Image 4 Source", "MK4"),
    ("image_5", "Image 5 Source", "MK5"),
    ("image_6", "Image 6 Source", "MK6"),
    ("image_7", "Image 7 Source", ""),
    ("image_8", "Image 8 Source", ""),
    ("image_9", "Image 9 Source", ""),
    ("variation_image", "Variation Image Source", ""),
]

IMAGE_MIX_ALIASES = {
    "SKUWA": ["SKUWA", "Seller SKU", "seller_sku"],
    "COLOR": ["COLOR", "Color", "Primary variation value", "Primary variation value (option)"],
    "SIZE": ["SIZE", "Size", "Secondary variation value", "Secondary variation value (option)"],
    "main_image": ["Main Image Source", "Main", "Main Image", "main_image"],
    "image_2": ["Image 2 Source", "Image 2", "Add 1", "Add1", "image_2"],
    "image_3": ["Image 3 Source", "Image 3", "Add 2", "Add2", "image_3"],
    "image_4": ["Image 4 Source", "Image 4", "Add 3", "Add3", "image_4"],
    "image_5": ["Image 5 Source", "Image 5", "Add 4", "Add4", "image_5"],
    "image_6": ["Image 6 Source", "Image 6", "Add 5", "Add5", "image_6"],
    "image_7": ["Image 7 Source", "Image 7", "Add 6", "Add6", "image_7"],
    "image_8": ["Image 8 Source", "Image 8", "Product Image 8", "Add 7", "Add7", "image_8"],
    "image_9": ["Image 9 Source", "Image 9", "Product Image 9", "image_9"],
    "variation_image": ["Variation Image Source", "Variation Image", "Primary variation image 1", "Swatch"],
}

SELLER_COLUMN_ALIASES = {
    "identifier_code_type": "gtin_type",
    "identifier_code": "gtin_code",
    "primary_variation_name": "property_name_1",
    "primary_variation_value": "property_value_1",
    "primary_variation_image_1": "property_1_image",
    "primary_variation_image_2": "property_1_image_2",
    "primary_variation_image_3": "property_1_image_3",
    "primary_variation_image_4": "property_1_image_4",
    "primary_variation_image_5": "property_1_image_5",
    "primary_variation_image_6": "property_1_image_6",
    "primary_variation_image_7": "property_1_image_7",
    "primary_variation_image_8": "property_1_image_8",
    "primary_variation_image_9": "property_1_image_9",
    "secondary_variation_name": "property_name_2",
    "secondary_variation_value": "property_value_2",
    "package_weight": "parcel_weight",
    "package_length": "parcel_length",
    "package_width": "parcel_width",
    "package_height": "parcel_height",
    "delivery_options": "delivery",
    "retail_price": "price",
    "auction_product": "special_product_listing_type",
    "starting_bid": "auction_starting_price",
    "materials": "product_property/100157",
    "pattern": "product_property/100198",
    "neckline": "product_property/100393",
    "sleeve_length": "product_property/100395",
    "sleeve_type": "product_property/100396",
    "season": "product_property/100397",
    "style": "product_property/100398",
    "fit": "product_property/100399",
    "stretch": "product_property/100400",
    "washing_instructions": "product_property/100401",
    "waist_height": "product_property/100403",
    "dangerous_goods": "product_property/101619",
    "organic_textile": "product_property/102418",
    "prop65_repro": "product_property/101395",
    "reprotoxic_chemicals": "product_property/101398",
    "prop65_carcinogens": "product_property/101400",
    "carcinogen": "product_property/101397",
    "sds": "qualification/8647636475739801353",
    "product_status": "aimed_product_status",
}


def get_profile_options():
    return [(key, profile["label"]) for key, profile in PRODUCT_PROFILES.items()]


def get_profile(profile_key):
    key = clean_text(profile_key).lower()
    if key not in PRODUCT_PROFILES:
        raise ValueError(f"Unsupported TikTok product profile: {profile_key}")
    return PRODUCT_PROFILES[key]


def build_tiktok_sku(sku_prefix, color, size_def):
    prefix = clean_text(sku_prefix)
    return f"{prefix}{color['sku_root']}{color['sku_letter']}{size_def['sku_code']}"


def generate_variants(profile_key, sku_prefix):
    profile = get_profile(profile_key)
    rows = []
    for size_def in SIZE_DEFINITIONS:
        for color in profile["colors"]:
            rows.append(
                {
                    "product_type": profile["label"],
                    "style": "5000",
                    "fulfillment": "Swift POD",
                    "size_input": size_def["input"],
                    "size_label": size_def["label"],
                    "size_code": size_def["sku_code"],
                    "color": color["name"],
                    "color_rule": color,
                    "sku": build_tiktok_sku(sku_prefix, color, size_def),
                }
            )
    return rows


def create_tiktok_update_template(output_path, profile_key="mens_tshirt", sku_prefix=""):
    profile = get_profile(profile_key)
    variants = []
    for color in profile["colors"]:
        for size_def in SIZE_DEFINITIONS:
            variants.append(
                {
                    "product_type": profile["label"],
                    "style": "5000",
                    "fulfillment": "Swift POD",
                    "size_input": size_def["input"],
                    "size_label": size_def["label"],
                    "size_code": size_def["sku_code"],
                    "color": color["name"],
                    "color_rule": color,
                    "sku": build_tiktok_sku(sku_prefix, color, size_def),
                }
            )
    rows = []

    for variant in variants:
        rows.append(
            {
                "Product Type": variant["product_type"],
                "Style": variant["style"],
                "Fulfillment": variant["fulfillment"],
                "Size": variant["size_input"],
                "TikTok Size": variant["size_label"],
                "Color": variant["color"],
                "SKU FF": f"{variant['color_rule']['sku_root']}{variant['color_rule']['sku_letter']}{variant['size_code']}",
                "Seller SKU": variant["sku"],
                "Brand": TIKTOK_DEFAULTS["brand"],
                "Product Name": "",
                "Product Description": "",
                "Main Image": "",
                "Image 8": "",
                "Image 9": "",
                "Price": TIKTOK_DEFAULTS["price"],
                "Quantity": TIKTOK_DEFAULTS["quantity"],
                "Size Chart": TIKTOK_DEFAULTS["size_chart"],
            }
        )

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False)
    return output_path


def create_tiktok_image_mixing_template(output_path, profile_key="mens_tshirt", sku_prefix="", randomize=False):
    profile = get_profile(profile_key)
    rows = []

    for color in profile["colors"]:
        row = {
            "Product Type": profile["label"],
            "Color": color["name"],
        }

        shuffled_codes = ["MK3", "MK4", "MK5", "MK6"]
        if randomize:
            random.shuffle(shuffled_codes)
        shuffled_by_key = {
            "image_3": shuffled_codes[0],
            "image_4": shuffled_codes[1],
            "image_5": shuffled_codes[2],
            "image_6": shuffled_codes[3],
        }

        for key, label, default_code in IMAGE_MIX_OUTPUTS:
            row[label] = shuffled_by_key.get(key, default_code)
        rows.append(row)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    pd.DataFrame(rows).to_excel(output_path, index=False)
    return output_path


def normalize_size(value):
    text = clean_text(value).upper().replace(" ", "")
    mapping = {
        "S": "S",
        "SMALL": "S",
        "M": "M",
        "MEDIUM": "M",
        "L": "L",
        "LARGE": "L",
        "XL": "XL",
        "XLARGE": "XL",
        "X-LARGE": "XL",
        "2XL": "2XL",
        "XXL": "2XL",
        "XXLARGE": "2XL",
        "XXLARGE": "2XL",
        "3XL": "3XL",
        "XXXL": "3XL",
        "4XL": "4XL",
        "XXXXL": "4XL",
        "5XL": "5XL",
        "XXXXXL": "5XL",
    }
    return mapping.get(text, text)


def normalize_color(value):
    return normalize_asset_key(value)


def fill_empty_cells(df):
    return df.where(pd.notna(df), "")


def load_tiktok_workbook(path, **kwargs):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported.*")
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")
        return load_workbook(path, **kwargs)


def read_update_file(update_file_path):
    if not update_file_path:
        return []

    ext = os.path.splitext(update_file_path)[1].lower()
    if ext == ".csv":
        df = fill_empty_cells(pd.read_csv(update_file_path, dtype=object, keep_default_na=False))
    elif ext in [".xlsx", ".xlsm", ".xls"]:
        df = fill_empty_cells(pd.read_excel(update_file_path, dtype=object, keep_default_na=False))
    else:
        raise ValueError(f"Unsupported update file type: {ext}")

    alias_lookup = {
        clean_text(alias).lower(): standard
        for standard, aliases in UPDATE_ALIASES.items()
        for alias in aliases
    }
    rename = {}
    for col in df.columns:
        standard = alias_lookup.get(clean_text(col).lower())
        if standard:
            rename[col] = standard
    df = df.rename(columns=rename)

    return [row.to_dict() for _, row in df.iterrows()]


def read_mapping_file(mapping_path):
    if not mapping_path:
        return []

    ext = os.path.splitext(mapping_path)[1].lower()
    if ext == ".csv":
        df = fill_empty_cells(pd.read_csv(mapping_path, dtype=object, keep_default_na=False))
    elif ext in [".xlsx", ".xlsm", ".xls"]:
        df = fill_empty_cells(pd.read_excel(mapping_path, sheet_name="Mapping", dtype=object, keep_default_na=False))
    else:
        raise ValueError(f"Unsupported mapping file type: {ext}")

    return [row.to_dict() for _, row in df.iterrows() if is_enabled(row.get("Enabled"))]


def read_image_mixing_file(image_mixing_path):
    if not image_mixing_path:
        return []

    ext = os.path.splitext(image_mixing_path)[1].lower()
    if ext == ".csv":
        df = fill_empty_cells(pd.read_csv(image_mixing_path, dtype=object, keep_default_na=False))
    elif ext in [".xlsx", ".xlsm", ".xls"]:
        df = fill_empty_cells(pd.read_excel(image_mixing_path, dtype=object, keep_default_na=False))
    else:
        raise ValueError(f"Unsupported image mixing file type: {ext}")

    alias_lookup = {
        clean_text(alias).lower(): standard
        for standard, aliases in IMAGE_MIX_ALIASES.items()
        for alias in aliases
    }
    rename = {}
    for col in df.columns:
        standard = alias_lookup.get(clean_text(col).lower())
        if standard:
            rename[col] = standard
    df = df.rename(columns=rename)

    return [row.to_dict() for _, row in df.iterrows()]


class UpdateLookup:
    def __init__(self, update_rows):
        self.rows = update_rows or []
        self.by_sku = {}
        self.by_color_size = {}

        for row in self.rows:
            sku = clean_text(row.get("SKUWA"))
            if sku:
                self.by_sku[sku] = row

            color = normalize_color(row.get("COLOR"))
            size = normalize_size(row.get("SIZE"))
            if color and size:
                self.by_color_size[(color, size)] = row

    def find(self, variant):
        sku = clean_text(variant["sku"])
        if sku in self.by_sku:
            return self.by_sku[sku]

        color_key = normalize_color(variant["color"])
        size_key = normalize_size(variant["size_input"])
        if (color_key, size_key) in self.by_color_size:
            return self.by_color_size[(color_key, size_key)]

        alias_keys = {normalize_color(alias) for alias in color_aliases(variant["color_rule"])}
        for (row_color, row_size), row in self.by_color_size.items():
            if row_size == size_key and row_color in alias_keys:
                return row

        if len(self.rows) == 1:
            return self.rows[0]

        return {}


class ImageMixLookup:
    def __init__(self, image_mix_rows):
        self.rows = image_mix_rows or []
        self.by_sku = {}
        self.by_color_size = {}
        self.by_color = {}

        for row in self.rows:
            sku = clean_text(row.get("SKUWA"))
            if sku:
                self.by_sku[sku] = row

            color = normalize_color(row.get("COLOR"))
            size = normalize_size(row.get("SIZE"))
            if color and size:
                self.by_color_size[(color, size)] = row
            if color and color not in self.by_color:
                self.by_color[color] = row

    def find(self, variant):
        sku = clean_text(variant["sku"])
        if sku in self.by_sku:
            return self.by_sku[sku]

        color_key = normalize_color(variant["color"])
        size_key = normalize_size(variant["size_input"])
        if (color_key, size_key) in self.by_color_size:
            return self.by_color_size[(color_key, size_key)]

        alias_keys = {normalize_color(alias) for alias in color_aliases(variant["color_rule"])}
        for alias_key in alias_keys:
            if (alias_key, size_key) in self.by_color_size:
                return self.by_color_size[(alias_key, size_key)]
            if alias_key in self.by_color:
                return self.by_color[alias_key]

        if len(self.rows) == 1:
            return self.rows[0]

        return {}


def first_non_empty(*values):
    for value in values:
        text = clean_text(value)
        if text:
            return value
    return ""


def parse_number(value, fallback):
    text = clean_text(value)
    if not text:
        return fallback
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return value


def format_mapping_value(value, expected_type):
    expected = clean_text(expected_type).lower()
    if expected == "number":
        return parse_number(value, "")
    return clean_text(value)


def build_label_map(ws):
    labels = {}
    for col in range(1, ws.max_column + 1):
        label = clean_text(ws.cell(row=3, column=col).value)
        if label and label not in labels:
            labels[label] = col
    return labels


def canonical_key(value):
    return normalize_asset_key(value)


def build_header_by_col(headers):
    return {col: key for key, col in headers.items()}


def build_alias_lookup(alias_map):
    return {
        clean_text(alias).lower(): standard
        for standard, aliases in alias_map.items()
        for alias in aliases
    }


def row_value(row, source_value, alias_map=None):
    if not row:
        return ""

    source = clean_text(source_value)
    if not source:
        return ""

    if source in row:
        return row.get(source)

    source_lower = source.lower()
    for key, value in row.items():
        if clean_text(key).lower() == source_lower:
            return value

    if alias_map:
        standard = build_alias_lookup(alias_map).get(source_lower)
        if standard and standard in row:
            return row.get(standard)

    source_norm = canonical_key(source)
    for key, value in row.items():
        if canonical_key(key) == source_norm:
            return value

    return ""


def payload_value(payload, source_value, target_key):
    candidates = [source_value, target_key]
    legacy_target = SELLER_COLUMN_ALIASES.get(clean_text(target_key).lower())
    if legacy_target:
        candidates.append(legacy_target)

    for candidate in candidates:
        key = clean_text(candidate)
        if not key:
            continue
        if key in payload:
            return payload.get(key)
        lower_key = key.lower()
        for payload_key, value in payload.items():
            if clean_text(payload_key).lower() == lower_key:
                return value
        norm_key = canonical_key(key)
        for payload_key, value in payload.items():
            if canonical_key(payload_key) == norm_key:
                return value

    return ""


def resolve_mapping_column(mapping_row, headers, label_map):
    seller_column = clean_text(mapping_row.get("Seller Column"))
    seller_columns = [seller_column]
    legacy_column = SELLER_COLUMN_ALIASES.get(seller_column.lower())
    if legacy_column:
        seller_columns.append(legacy_column)

    for column_name in seller_columns:
        if column_name in headers:
            return headers[column_name]
        if column_name in label_map:
            return label_map[column_name]

    seller_excel_col = clean_text(mapping_row.get("Seller Excel Col"))
    if seller_excel_col:
        try:
            return column_index_from_string(seller_excel_col)
        except Exception:
            return None

    return None


def resolve_getlinks_mapping_value(mapping_row, target_key, context):
    source_value = clean_text(mapping_row.get("Source Value"))
    payload = context["payload"]

    if looks_like_url(source_value):
        return source_value

    if not source_value:
        return payload_value(payload, "", target_key)

    existing = payload_value(payload, source_value, target_key)
    if clean_text(existing) and canonical_key(source_value) not in {
        "mk1",
        "mk2",
        "mk3",
        "mk4",
        "mk5",
        "mk6",
        "mk7",
        "mk8",
        "mk9",
        "mk",
    }:
        return existing

    resolver = context["resolver"]
    color_rule = context["variant"]["color_rule"]

    if target_key == "main_image" and canonical_key(source_value).startswith("mk1"):
        return resolver.get_url(
            main_image_candidates(color_rule),
            label=f"main image for {context['variant']['color']}",
            loose_tokens=["mk1", color_rule["name"]],
            required=True,
        )

    return resolve_source_code(
        resolver,
        source_value,
        color_rule,
        label=f"{source_value} for {context['variant']['color']}",
    )


def resolve_mapping_value(mapping_row, target_key, context):
    source_type = clean_text(mapping_row.get("Source Type")).lower()
    source_value = clean_text(mapping_row.get("Source Value"))

    if source_type in ["skip", ""]:
        return None, False

    if source_type == "fixed":
        return mapping_row.get("Fixed Value"), True

    if source_type == "blank":
        return "", True

    if source_type in ["auto", "generated", "payload"]:
        return payload_value(context["payload"], source_value, target_key), True

    if source_type == "update_file":
        value = row_value(context["update_row"], source_value, UPDATE_ALIASES)
        if clean_text(value):
            return value, True

        if not target_key.startswith("image_") and target_key != "main_image":
            return payload_value(context["payload"], source_value, target_key), True

        return "", True

    if source_type == "getlinks":
        return resolve_getlinks_mapping_value(mapping_row, target_key, context), True

    return None, False


def apply_mapping_overrides(ws, headers, label_map, row, mapping_rows, context):
    header_by_col = build_header_by_col(headers)

    for mapping_row in mapping_rows:
        source_type = clean_text(mapping_row.get("Source Type")).lower()
        if source_type == "skip":
            continue

        col = resolve_mapping_column(mapping_row, headers, label_map)
        if not col or col > ws.max_column:
            continue

        target_key = header_by_col.get(col, clean_text(mapping_row.get("Seller Column")))
        legacy_target = SELLER_COLUMN_ALIASES.get(target_key.lower())
        if legacy_target:
            target_key = legacy_target

        value, should_write = resolve_mapping_value(mapping_row, target_key, context)
        if not should_write:
            continue

        ws.cell(row=row, column=col).value = format_mapping_value(
            value,
            mapping_row.get("Expected Type"),
        )


def color_aliases(color_rule):
    values = [color_rule["name"], color_rule["image_suffix"]]
    values.extend(color_rule.get("aliases", []))
    return [clean_text(value) for value in values if clean_text(value)]


def image_candidates(code, color_rule=None):
    base = clean_text(code)
    candidates = []

    if color_rule:
        for alias in color_aliases(color_rule):
            candidates.extend(
                [
                    f"{base}-{alias}",
                    f"{base}_{alias}",
                    f"{base} {alias}",
                    f"{base}{alias}",
                ]
            )

    candidates.append(base)
    return candidates


def main_image_candidates(color_rule):
    candidates = []
    for alias in color_aliases(color_rule):
        candidates.extend(
            [
                f"MK1-{alias}-3x4",
                f"MK1_{alias}_3x4",
                f"MK1 {alias} 3x4",
                f"MK1{alias}3x4",
            ]
        )
    candidates.extend(image_candidates("MK1", color_rule))
    for alias in color_aliases(color_rule):
        candidates.extend(
            [
                f"main-{alias}",
                f"main_{alias}",
                f"main {alias}",
                f"main{alias}",
            ]
        )
    candidates.append("main")
    candidates.extend(color_aliases(color_rule))
    return candidates


def variation_image_candidates(color_rule):
    candidates = []
    for alias in color_aliases(color_rule):
        candidates.extend(
            [
                alias,
                f"color-{alias}",
                f"color_{alias}",
                f"swatch-{alias}",
                f"swatch_{alias}",
                f"sw-{alias}",
                f"sw_{alias}",
            ]
        )
    candidates.extend(image_candidates("MK1", color_rule))
    return candidates


def looks_like_url(value):
    text = clean_text(value).lower()
    return text.startswith("http://") or text.startswith("https://")


def image_source_candidates(source_code, color_rule):
    source_code = clean_text(source_code)
    candidates = image_candidates(source_code, color_rule)
    loose_tokens = [source_code]

    if canonical_key(source_code) == "mk2":
        generic_candidates = candidates[-1:]
        mk_candidates = []
        for alias in color_aliases(color_rule):
            if clean_text(alias).isdigit():
                continue
            mk_candidates.extend(
                [
                    f"MK-{alias}",
                    f"MK_{alias}",
                    f"MK {alias}",
                    f"MK{alias}",
                ]
            )
        candidates = candidates[:-1] + mk_candidates + generic_candidates

    suffixes = [clean_text(color_rule.get("image_suffix"))]
    suffixes.extend(color_aliases(color_rule))
    for suffix in [item for item in suffixes if item]:
        for separator in ["-", "_", " "]:
            suffix_token = f"{separator}{suffix}"
            if source_code.lower().endswith(suffix_token.lower()):
                base_code = source_code[: -len(suffix_token)]
                if base_code:
                    candidates.extend(image_candidates(base_code, color_rule))
                    loose_tokens.append(base_code)

    return candidates, loose_tokens


def mix_code_for(image_mix_row, output_key):
    value = clean_text(image_mix_row.get(output_key))
    if value:
        return value

    for key, _, default_code in IMAGE_MIX_OUTPUTS:
        if key == output_key:
            return default_code
    return ""


def resolve_source_code(resolver, source_code, color_rule, label="", required=False):
    source_code = clean_text(source_code)
    if not source_code:
        return ""

    if looks_like_url(source_code):
        return source_code

    candidates, loose_tokens = image_source_candidates(source_code, color_rule)
    return resolver.get_url(
        candidates,
        label=label or source_code,
        loose_tokens=loose_tokens,
        required=required,
    )


def resolve_images_for_variant(resolver, variant, update_row, image_mix_row):
    color_rule = variant["color_rule"]
    main_override = clean_text(update_row.get("MAIN_IMAGE"))
    image_9_override = clean_text(update_row.get("IMAGE_9"))

    if main_override:
        main_image = main_override
    else:
        main_code = mix_code_for(image_mix_row, "main_image")
        if canonical_key(main_code).startswith("mk1"):
            main_image = resolver.get_url(
                main_image_candidates(color_rule),
                label=f"main image for {variant['color']}",
                loose_tokens=["mk1", variant["color"]],
                required=False,
            )
        else:
            main_image = resolve_source_code(
                resolver,
                main_code,
                color_rule,
                label=f"{main_code} main image for {variant['color']}",
                required=False,
            )
        if not main_image:
            main_image = resolver.get_url(
                main_image_candidates(color_rule),
                label=f"main image for {variant['color']}",
                loose_tokens=["mk1", color_rule["image_suffix"]],
                required=True,
            )

    product_images = {}
    for header in ["image_2", "image_3", "image_4", "image_5", "image_6", "image_7", "image_8"]:
        update_override = clean_text(update_row.get(header.upper()))
        if update_override:
            product_images[header] = update_override
        else:
            code = mix_code_for(image_mix_row, header)
            product_images[header] = resolve_source_code(
                resolver,
                code,
                color_rule,
                label=f"{code} for {variant['color']}",
            )

    if image_9_override:
        product_images["image_9"] = image_9_override
    elif mix_code_for(image_mix_row, "image_9"):
        image_9_code = mix_code_for(image_mix_row, "image_9")
        product_images["image_9"] = resolve_source_code(
            resolver,
            image_9_code,
            color_rule,
            label=f"{image_9_code} for image 9",
        )
    else:
        product_images["image_9"] = resolver.get_url(
            ["Option_colors", "Option colors", "Color chart", "Colors", "Swatch", "SWATCH"],
            label="option colors / swatch",
            loose_tokens=["option", "color"],
        )

    variation_code = mix_code_for(image_mix_row, "variation_image")
    variation_image = resolve_source_code(
        resolver,
        variation_code,
        color_rule,
        label=f"{variation_code} variation image for {variant['color']}",
    )
    if not variation_image:
        variation_image = resolver.get_url(
            variation_image_candidates(color_rule),
            label=f"variation image for {variant['color']}",
            loose_tokens=[color_rule["image_suffix"]],
        )

    if not variation_image:
        variation_image = main_image

    return main_image, product_images, variation_image


def build_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        key = clean_text(ws.cell(row=1, column=col).value)
        if key and key not in headers:
            headers[key] = col
    return headers


def clear_data_rows(ws, start_row=DATA_START_ROW):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def copy_row_style(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)


def set_header_value(ws, headers, row, key, value):
    col = headers.get(key)
    if not col:
        return False

    cell = ws.cell(row=row, column=col)
    if isinstance(value, str):
        cell.number_format = "@"
    cell.value = value
    return True


def build_row_payload(profile, variant, update_row, image_mix_row, resolver, params):
    product_name = first_non_empty(
        update_row.get("PRODUCT_NAME"),
        params.get("product_name"),
        profile["default_name"],
    )
    product_description = first_non_empty(
        update_row.get("PRODUCT_DESCRIPTION"),
        params.get("product_description"),
        DEFAULT_DESCRIPTION,
    )
    brand = first_non_empty(update_row.get("BRAND"), params.get("brand"), TIKTOK_DEFAULTS["brand"])
    price = parse_number(first_non_empty(update_row.get("PRICE"), params.get("price")), TIKTOK_DEFAULTS["price"])
    quantity = parse_number(first_non_empty(update_row.get("QUANTITY"), params.get("quantity")), TIKTOK_DEFAULTS["quantity"])
    size_chart = first_non_empty(update_row.get("SIZE_CHART"), params.get("size_chart"), TIKTOK_DEFAULTS["size_chart"])

    main_image, product_images, variation_image = resolve_images_for_variant(
        resolver,
        variant,
        update_row,
        image_mix_row,
    )

    payload = {
        "category": profile["category"],
        "brand": brand,
        "product_name": product_name,
        "product_description": product_description,
        "main_image": main_image,
        "property_name_1": "Color",
        "property_value_1": variant["color"],
        "property_1_image": variation_image,
        "property_name_2": "Size",
        "property_value_2": variant["size_label"],
        "price": price,
        "quantity": quantity,
        "seller_sku": variant["sku"],
        "size_chart": size_chart,
    }

    payload.update(product_images)
    payload.update(TIKTOK_DEFAULTS)
    payload["brand"] = brand
    payload["price"] = price
    payload["quantity"] = quantity
    payload["size_chart"] = size_chart
    payload["seller_sku"] = variant["sku"]
    payload["main_image"] = main_image

    return payload


def generate_tiktok_upload_file(
    template_path,
    image_folder,
    output_folder,
    sku_prefix,
    profile_key="mens_tshirt",
    cloud_folder="tiktok",
    env_path=".env",
    product_name="",
    product_description="",
    brand="No brand",
    price=19.99,
    quantity=100,
    size_chart="7660455797738030862",
    update_file_path="",
    image_mixing_path="",
    mapping_path="",
    upload_images=True,
    log_callback=None,
):
    if not os.path.exists(template_path):
        raise ValueError(f"TikTok template not found: {template_path}")

    if not clean_text(sku_prefix):
        raise ValueError("SKU Prefix is required.")

    os.makedirs(output_folder, exist_ok=True)

    profile = get_profile(profile_key)
    variants = generate_variants(profile_key, sku_prefix)
    update_lookup = UpdateLookup(read_update_file(update_file_path))
    image_mix_lookup = ImageMixLookup(read_image_mixing_file(image_mixing_path))
    mapping_rows = read_mapping_file(mapping_path)

    resolver = ImageResolver(
        image_folder=image_folder,
        cloud_folder=cloud_folder,
        env_path=env_path,
        upload_enabled=upload_images,
        log_callback=log_callback,
    )

    if log_callback:
        log_callback(f"Images found: {resolver.image_count}")
        log_callback(f"Variants generated: {len(variants)}")
        if image_mixing_path:
            log_callback(f"Image mixing rows: {len(image_mix_lookup.rows)}")
        if mapping_path:
            log_callback(f"Mapping rows: {len(mapping_rows)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = os.path.join(output_folder, f"final_tiktok_upload_ready_{timestamp}.xlsx")
    log_csv = os.path.join(output_folder, f"tiktok_generation_log_{timestamp}.csv")

    shutil.copy2(template_path, output_xlsx)
    wb = load_tiktok_workbook(output_xlsx)

    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TEMPLATE_SHEET}' not found in TikTok template. Available sheets: {wb.sheetnames}")

    ws = wb[TEMPLATE_SHEET]
    headers = build_header_map(ws)
    label_map = build_label_map(ws)
    clear_data_rows(ws)

    thin_border = Border(
        left=Side(style="thin", color="D4D4D4"),
        right=Side(style="thin", color="D4D4D4"),
        top=Side(style="thin", color="D4D4D4"),
        bottom=Side(style="thin", color="D4D4D4"),
    )

    logs = []
    missing_main = []
    params = {
        "product_name": product_name,
        "product_description": product_description,
        "brand": brand,
        "price": price,
        "quantity": quantity,
        "size_chart": size_chart,
    }

    for index, variant in enumerate(variants):
        excel_row = DATA_START_ROW + index
        copy_row_style(ws, DATA_START_ROW, excel_row)

        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).border = thin_border

        update_row = update_lookup.find(variant)
        image_mix_row = image_mix_lookup.find(variant)
        payload = build_row_payload(profile, variant, update_row, image_mix_row, resolver, params)

        if not clean_text(payload.get("main_image")):
            missing_main.append(f"{variant['color']} {variant['size_input']}")

        for key, value in payload.items():
            set_header_value(ws, headers, excel_row, key, value)

        apply_mapping_overrides(
            ws,
            headers,
            label_map,
            excel_row,
            mapping_rows,
            {
                "payload": payload,
                "update_row": update_row,
                "image_mix_row": image_mix_row,
                "variant": variant,
                "resolver": resolver,
            },
        )

        logs.append(
            {
                "Excel Row": excel_row,
                "Product Type": variant["product_type"],
                "Color": variant["color"],
                "Size": variant["size_input"],
                "TikTok Size": variant["size_label"],
                "Seller SKU": variant["sku"],
                "Main Image": payload.get("main_image", ""),
                "Variation Image": payload.get("property_1_image", ""),
            }
        )

    if missing_main:
        sample = ", ".join(missing_main[:8])
        raise ValueError(f"Missing mandatory main image for variants: {sample}")

    wb.save(output_xlsx)

    with open(log_csv, "w", newline="", encoding="utf-8-sig") as file:
        fieldnames = [
            "Excel Row",
            "Product Type",
            "Color",
            "Size",
            "TikTok Size",
            "Seller SKU",
            "Main Image",
            "Variation Image",
        ]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(logs)

    return output_xlsx, log_csv
